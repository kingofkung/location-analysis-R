##---------------------------------------------------------------------
## Program: roadDistanceFinder.py
## Version: 2.1
##
## Author: Nick Joslyn
## Institution: Simpson College
##
## Acknowledgements: Professor Paul Craven (Simpson College) for introducing
##	me to the Python Google Maps API. Advisor Heidi Berger and research 
##	partners Emma Christensen and Maddy Kersten.
##
## Purpose: 
##	This program is written to extract Google Maps' recommended
##	route distance between a set of demand nodes and one facility
##	node. This program reads in a set of voter addresses from an Excel 
##	file, calculates the distances to one facility node, and writes out 
##	the distances and times to a specified Excel file. The output Excel 
##	file will contain two columns of distances. Column A will contain 
##	the travel distance in meters. Column B will contain the travel 
##	duration in seconds.
##
## Instructions:
##	This program is designed as a black box. Everything is generalized
##	and runs off of user-input. If you desire, you can edit the
##	code. it is thoroughly commented.
##
##	You must have openpyxl and the Google Maps modules installed.
##	Additionally, you must have a Google Maps API key. API keys have
##	quotas. 2500 requests per day. Only 100 at a time. If you enable
##	billing, your quotas increase, and you should increase the apiSingleQuota
##	variable accordingly.
##
##	This is a Python script. You can run it using your preferred Python
##	editor -- i.e. IDLE, SUBLIME, VIM, etc.
##	
##	For SUBLIME Text, you must have the REPL installed. Otherwise, 
##	the input arguments will not be processesed. If using SUBLIME, run
##	using the REPL Python build.
##
## Notes: 
##	If Google cannot find the address, the code is written to put a
##	blank cell in the output file. If you find a blank cell in your
##	output file, then you must manually decide how to deal with that
##	particular data point. This will not happen often, and the program
##	will tell you which row was left blank.
##	
##	Google Maps is a dynamic service. Occasionally, the same demand
##	and facility nodes will produce different results depending on the 
##	optimized route chosen by Google.
##---------------------------------------------------------------------

##=====================================================================

## This section uses user input to determine the six important variables
## that instruct the rest of the program on how to run.

# Input is an Excel file with demand addresses
inputFileName = input("Type the name of the demand node Excel file (Specify .xlsx extension.): ")

# This allows the program to skip over any header information rows
headerOffset = int(input("What row does your data start on? (Specify numerical value. Likely 2.): "))

# This indicates to the program which column in the Excel file contains the addresses
columnAddress = int(input("What column number are the demand node addresses in? (Specify numerical value.): "))

# This is the address of the single facility node
facilityAddress = input("What is the address of the facility node to find the distances for?: ")

# This is the API Key of the user to allow Google Maps information extraction
apiKey = input("What is your Google Maps API Key?: ")

# This is the Excel file where the distances and times will be written out to
outputFileName = input("Type the name of an Excel file you want the results printed to (Specify .xlsx extension. New file recommended): ")

# This is not a user input driven variable.
# However, it is included in this section because it could be changed.
# If you enabled billing, this could be much higher.
# Without enabling billing, 100 is the maximum amount allowed by Google in one call.
apiSingleQuota = 100

##=====================================================================

## From here on, the program runs in a self-contained manner.
## Everything is based on the user input variables.

print("\n----------------Program Running----------------\n")

# Module for reading/writing from Excel Files
# Module for timing how long the program takes
# Modules for extracting Google Maps information

import openpyxl
from openpyxl import Workbook

import time

import googlemaps
from datetime import datetime

# Begin Timer
start = time.time()

# Load the file we want to read from and set sheet to hold the information.
book = openpyxl.load_workbook(inputFileName)
sheet = book.active

# Count the number of demands that need to go to a facility.
# We subtract the headerOffset (minus 1) to subtract out rows 
# that are not demand nodes.
numberOfDemandNodes = sheet.max_row - (headerOffset - 1)

# Loop through the spreadsheet, filling up demandAddressList with
# the address of each demand node in sequential order.
# Note, we add the headerOffset variable to the loop so we start with the correct row
demandAddressList = []
for i in range(numberOfDemandNodes):
		demandAddressList.append(sheet.cell(row = (i+headerOffset), column = columnAddress).value)

# Set the active worksheet to begin the write-out process
wb = Workbook()
ws = wb.active

# Supply header information for the two columns that we will write
# to in the output file
ws.cell(row = 1, column = 1, value = "Distance (meters)")
ws.cell(row = 1, column = 2, value = "Time (seconds)")

# The counter variable will be updated as an index for the output file while we loop 
# through the demand nodes. It needs to be offset by two to begin after the header information
counter = 0

# We will iterate through every demand node, find the distance between
# the address and the facility node, and put that value in
# the spreadsheet (meters and seconds). Loop by the maximum allowable number of
# elements in a single Google Maps call (apiSingleQuota) for speed.
for i in range(numberOfDemandNodes//apiSingleQuota):

	# Initialize the connection to Google Maps
	gmaps = googlemaps.Client(key = apiKey)

	# Extract all of the information provieded by Google
	result = gmaps.distance_matrix(demandAddressList[counter:(counter + apiSingleQuota)], facilityAddress)

	# Grab the road travel distance in meters and put it in the
	# spreadsheet at the correct index. If the GoogleMaps
	# fails, nothing will be put into the Excel sheet. Thus, find the demand 
	# node who needs to travel an empty distance and manually decide what to do with the data.
	for row in result['rows']:
		for column in row['elements']:
			try:
				ws.cell(row = (counter + 2), column = 1, value = column['distance']['value'])
				ws.cell(row = (counter + 2), column = 2, value = column['duration']['value'])
			except:
				print("Error. The value could not be found. Row: " + str(counter + 2))
		
		# Update the counter	
		counter = counter + 1

##--------------
## Run again for the remaining addresses that were not a perfect multiple of apiSingleQuota
##--------------

# Initialize the connection to Google Maps
gmaps = googlemaps.Client(key = apiKey)

# Extract all of the information provieded by Google
result = gmaps.distance_matrix(demandAddressList[counter:], facilityAddress)

# Grab the road travel distance in meters and put it in the
# spreadsheet at the correct index. If the GoogleMaps
# fails, nothing will be put into the Excel sheet. Thus, find the demand 
# node who needs to travel an empty distance and manually decide what to do with the data.
for row in result['rows']:
	for column in row['elements']:
		try:
			ws.cell(row = (counter + 2), column = 1, value = column['distance']['value'])
			ws.cell(row = (counter + 2), column = 2, value = column['duration']['value'])

		except:
			print("Error. The value could not be found. Row: " + str(counter + 2))

		counter = counter + 1

# Save the output Excel file with the user-inputted name
wb.save(outputFileName)

print("\n----------------Program Done----------------\n")

print("Results written to " + str(outputFileName) + "\n")

# End Timer
# Print elapsed time rounded to two places
end = time.time()
print("Time for program to complete: " + str(round(end - start, 2)))
